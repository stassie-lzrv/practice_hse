import csv
import openpyxl
import pandas as pd
def append_csv_to_excel(csv_path, xlsx_path):
    # Открываем CSV файл для чтения
    with open(csv_path, 'r') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=';')

        # Открываем XLSX файл для записи
        wb = openpyxl.load_workbook(xlsx_path)
        sheet = wb.active

        # Получаем максимальное количество столбцов в XLSX файле
        max_columns = sheet.max_column

        # Получаем последнюю заполненную строку в XLSX файле
        last_row = sheet.max_row + 1

        # Перебираем строки в CSV файле
        for row in csv_reader:
            # Перебираем значения в каждой строке CSV файла
            for i, value in enumerate(row):
                # Добавляем значение в соответствующую ячейку XLSX файла
                sheet.cell(row=last_row, column=i + 1).value = value

            # Увеличиваем значение последней строки XLSX файла
            last_row += 1

        # Сохраняем изменения в XLSX файле
        wb.save(xlsx_path)


def calculate_data(file_path):
    # Чтение xlsx файла
    df = pd.read_excel(file_path)

    # Подсчет процента строк, содержащих "Да" в столбце 21
    yes_count = df[df.iloc[:, 20] == "Да"].shape[0]
    total_count = df.shape[0]
    percentage = (yes_count / total_count) * 100

    return percentage


def create_new_file(file_path):
    # Чтение xlsx файла
    df = pd.read_excel(file_path)

    # Создание нового xlsx файла с уникальными значениями из столбца 41
    new_df = pd.DataFrame(df.iloc[:, 40].unique(), columns=["Наименование учебного заведения"])
    new_file_path = "schools.xlsx"
    new_df.to_excel(new_file_path, index=False)

    return new_file_path


csv_path = input("Введите путь к CSV файлу: ")
xlsx_path = input("Введите путь к XLSX файлу: ")

try:
    append_csv_to_excel(csv_path, xlsx_path)
    print("Новые данные успешно добавлены в XLSX файл")
    percentage = calculate_data(xlsx_path)
    new_file_path = create_new_file(xlsx_path)
    print("Процент абитуриентов, достигших статуса студент ФКН: ", percentage)
    print("Файл, содержащий наименования учебных заведений:", new_file_path)
except Exception as e:
    print(f"Произошла ошибка: {e}")